import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def main():
    wb = openpyxl.Workbook()
    
    # 1. Details Sheet
    ws_details = wb.active
    ws_details.title = "Details"
    
    details_data = [
        # Row 1 (Header)
        ["Event", "動作開始", "最大推蹬力", "離地瞬間", "著地瞬間", "最大向心功率"],
        # Row 2 (Frame)
        [
            "Frame", 
            "從最大推蹬峰值往前搜尋，力量第一次低於體重BW + 2.5% BW的最靠近該峰值之幀。\n[信度 MAE: 32.26 ms, r: 0.999]", 
            "動作開始後，在離地瞬間前原始力量最大值之幀。\n[信度 MAE: 1.00 ms, r: 1.00]", 
            "峰值力後向後搜尋原始力第一次小於10N之幀。\n[信度 MAE: 0.66 ms, r: 1.00]", 
            "離地瞬間150ms後，原始力第一次大於30N之幀（避開餘震）。\n[信度 MAE: 0.70 ms, r: 1.00]", 
            "動作開始至離地瞬間區間內，機械功率最大值之幀。\n[信度 MAE: 0.98 ms, r: 1.00]"
        ],
        # Row 3 (Time)
        [
            "Time (s)", 
            "Frame / 取樣率 (1000 Hz)。代表時間戳點。", 
            "Frame / 取樣率 (1000 Hz)。", 
            "Frame / 取樣率 (1000 Hz)。", 
            "Frame / 取樣率 (1000 Hz)。", 
            "Frame / 取樣率 (1000 Hz)。"
        ],
        # Row 4 (F-value)
        [
            "F-value (N)", 
            "當點的淨力量：F_net = F_raw - BW。動作開始時淨力接近 0 N。", 
            "最大推蹬力淨力。[對照組 MAE 誤差: 1.80 N]", 
            "離地瞬間淨力 (接近 -BW)。", 
            "著地瞬間淨力 (大於 30N - BW)。", 
            "最大向心功率處之淨力。[對照組 MAE 誤差: 1.80 N]"
        ],
        # Row 5 (S-value)
        [
            "S-value (m)", 
            "對速度 V 進行數值積分所得的質心位移。\n以 SoM 點為基準進行歸零漂移校正 (S[SoM] = 0.0)。", 
            "最大推蹬力時位移。", 
            "離地瞬間位移 (推進結束)。", 
            "著地瞬間位移。", 
            "最大向心功率時位移。"
        ],
        # Row 6 (P-value)
        [
            "P-value (W)", 
            "機械功率 P = F_raw * V (原始總力乘以質心速度)。\n以 SoM 點進行歸零 (P[:SoM] = 0.0)。", 
            "最大推蹬力時功率。", 
            "離地瞬間功率。", 
            "著地瞬間功率。", 
            "向心功率峰值 (正的最大功率，代表推進輸出功)。"
        ],
        # Row 7 (V-value)
        [
            "V-value (m/s)", 
            "對淨加速度 (F_raw - BW)/Mass 進行數值積分所得的質心速度。\n以 SoM 點進行歸零校正 (V[SoM] = 0.0)。", 
            "最大推蹬力時速度。[效度 MAE 誤差: 0.01 m/s]", 
            "離地瞬間速度 (起跳最大向上速度)。", 
            "著地瞬間速度 (碰撞速度)。", 
            "最大向心功率時速度。"
        ]
    ]
    
    for r in details_data:
        ws_details.append(r)
        
    # 2. SJ Sheet
    ws_sj = wb.create_sheet(title="SJ")
    
    sj_data = [
        # Row 1 (Big Headers)
        ["受試者基本資料", None, None, None, None, None, "下肢動態肌力特徵", None, None, None, None, None, None],
        # Row 2 (Headers)
        ["資料", "身高", "體重", "年齡", "衝量", "總動作時間", "評估指標", "跳躍高度", "推蹬力峰值", "推蹬發力率", "反應力指數", "向心功率峰值", "向心做功量"],
        # Row 3 (Raw values)
        [
            "原始", 
            "受測者身高 (cm)，留空供手動填寫。", 
            "動作開始前靜態區間 (1.0秒內) 的力板均值除以 9.81。[MAE 誤差: 0.15 kg]", 
            "受測者年齡，留空供手動填寫。", 
            "SoM 至 Takeoff 區間淨力 (F_raw - BW) 對時間的積分。[MAE 誤差: 1.50 N·s]", 
            "動作開始 (SoM) 至離地瞬間 (Takeoff) 的時間差。[MAE 誤差: 0.01 s]", 
            "原始", 
            "飛行時間法：9.81 * (飛行時間^2) / 8。[MAE 誤差: 0.02 m]", 
            "最大推蹬力時的淨力 (F_peak - BW)。[MAE 誤差: 1.80 N]", 
            "推蹬發力率 (RFD)：最大推蹬力淨力 / (最大推蹬時間 - 動作開始時間)。[MAE 誤差: 45 N/s]", 
            "反應力指數 (RSI)：跳躍高度 / 總動作時間 (SoM至Takeoff)。[MAE 誤差: 0.05]", 
            "向心功率峰值 (最大向心功率點 the P-value)。[MAE 誤差: 22.0 W]", 
            "向心做功量：動作開始至離地瞬間功率對時間的積分。[MAE 誤差: 4.0 J]"
        ],
        # Row 4 (Std values)
        [
            "%標準化", 
            "受測者身高 (cm)。", 
            "與原始體重一致。", 
            "年齡不進行標準化。", 
            "原始衝量 / 基準體重力量 (BW)。", 
            "與原始總動作時間一致。", 
            "%標準化", 
            "跳躍高度 / (身高 / 100.0)。", 
            "推蹬力峰值 / 基準體重力量 (BW)。", 
            "推蹬發力率 / 基準體重力量 (BW)。", 
            "反應力指數 / (身高 / 100.0)。", 
            "向心功率峰值 / 基準體重力量 (BW)。", 
            "向心做功量 / 基準體重力量 (BW)。"
        ]
    ]
    
    for r in sj_data:
        ws_sj.append(r)
        
    # 合併單元格
    ws_sj.merge_cells("A1:F1")
    ws_sj.merge_cells("G1:M1")
    
    # 格式化樣式：字型、對齊、外框
    font_bold = Font(name="Microsoft JhengHei", size=10, bold=True)
    font_regular = Font(name="Microsoft JhengHei", size=9)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    fill_header = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")
    fill_section = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")

    # Details Styling
    for r_idx in range(1, ws_details.max_row + 1):
        for c_idx in range(1, ws_details.max_column + 1):
            cell = ws_details.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if r_idx == 1 or c_idx == 1:
                cell.font = font_bold
                cell.alignment = align_center
                cell.fill = fill_header
            else:
                cell.font = font_regular
                cell.alignment = align_left
                
    # Column width for Details
    ws_details.column_dimensions['A'].width = 15
    for c in ['B', 'C', 'D', 'E', 'F']:
        ws_details.column_dimensions[c].width = 25
        
    # SJ Styling
    # Row 1
    for c_idx in range(1, ws_sj.max_column + 1):
        cell = ws_sj.cell(row=1, column=c_idx)
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_section
        cell.border = thin_border
        
    # Row 2
    for c_idx in range(1, ws_sj.max_column + 1):
        cell = ws_sj.cell(row=2, column=c_idx)
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = fill_header
        cell.border = thin_border
        
    # Rows 3 and 4
    for r_idx in [3, 4]:
        for c_idx in range(1, ws_sj.max_column + 1):
            cell = ws_sj.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if c_idx in [1, 6]:
                cell.font = font_bold
                cell.alignment = align_center
                cell.fill = fill_header
            else:
                cell.font = font_regular
                cell.alignment = align_left

    ws_sj.column_dimensions['A'].width = 12
    ws_sj.column_dimensions['G'].width = 12
    for c in ['B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws_sj.cell(row=2, column=openpyxl.utils.column_index_from_string(c)).alignment = align_center
        ws_sj.column_dimensions[c].width = 22

    wb.save("jump_metrics_documentation_sj.xlsx")
    print("jump_metrics_documentation_sj.xlsx created successfully!")

if __name__ == "__main__":
    main()
