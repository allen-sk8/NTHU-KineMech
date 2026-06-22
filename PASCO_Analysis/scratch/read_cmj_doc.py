import openpyxl

def main():
    path = "jump_metrics_documentation_cmj.xlsx"
    wb = openpyxl.load_workbook(path)
    print("Sheets in CMJ doc:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- Sheet: {name} (Rows: {ws.max_row}, Cols: {ws.max_column}) ---")
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [cell.value for cell in ws[r]]
            print(f"Row {r}: {row_vals[:15]}")
            
if __name__ == "__main__":
    main()
