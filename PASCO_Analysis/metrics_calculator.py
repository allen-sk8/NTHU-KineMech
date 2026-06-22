import numpy as np
import pandas as pd
import openpyxl

def calculate_advanced_metrics(event_vals, force_series, p_series, bw, fs, height=np.nan, age=np.nan):
    """
    計算 CMJ 垂直跳躍之進階動態肌力與時宜特徵，對齊實驗室金標準的第二工作表 (CMJ)

    參數:
        event_vals (dict): 包含 11 個事件的 Frame、Time (s)、F-value (N)、S-value (m) 等數值的字典
        force_series (np.ndarray): 原始力量序列
        p_series (np.ndarray): 功率序列
        bw (float): 基準體重 (N)
        fs (int): 取樣率 (Hz)
        height (float): 身高 (cm)，若無則設為 np.nan
        age (float): 年齡，若無則設為 np.nan
    """
    dt = 1.0 / fs
    bw_n = bw
    weight_kg = round(bw / 9.81, 2)
    
    # 提取四捨五入至兩位小數的特徵時間點 (同對照組時間差計算方式)
    t_som = round(event_vals["動作開始"]["Time"], 2)
    t_unweight = round(event_vals["最大失重"]["Time"], 2)
    t_braking = round(event_vals["制動開始"]["Time"], 2)
    t_lowest = round(event_vals["重心最低"]["Time"], 2)
    t_peak = round(event_vals["最大推蹬力"]["Time"], 2)
    t_takeoff = round(event_vals["離地瞬間"]["Time"], 2)
    t_landing = round(event_vals["著地瞬間"]["Time"], 2)
    
    t_amort_start = round(event_vals["攤還期開始"]["Time"], 2)
    t_amort_end = round(event_vals["攤還期結束"]["Time"], 2)
    
    # 動作索引
    som_idx = int(event_vals["動作開始"]["Frame"])
    takeoff_idx = int(event_vals["離地瞬間"]["Frame"])
    lowest_idx = int(event_vals["重心最低"]["Frame"])
    
    # --- 1. 受試者基本資料 ---
    # 衝量 (N·s)
    impulse_raw = np.sum((force_series[som_idx:takeoff_idx] - bw) * dt)
    impulse_std = impulse_raw / bw_n
    
    # --- 2. 下肢動態肌力特徵 ---
    # 跳躍高度 (飛行時間法)
    t_flight = t_landing - t_takeoff
    jump_height = 9.81 * (t_flight ** 2) / 8
    jump_height_std = jump_height / (height / 100.0) if not np.isnan(height) else np.nan
    
    # 推蹬力峰值 (淨力)
    peak_force_con = event_vals["最大推蹬力"]["F"]
    peak_force_con_std = peak_force_con / bw_n
    
    # 推蹬發力率 (RFD)
    rfd_con = peak_force_con / (t_peak - t_braking) if t_peak != t_braking else 0.0
    rfd_con_std = rfd_con / bw_n
    
    # 反應力指數 (RSI)
    total_action_time = t_takeoff - t_som
    rsi = jump_height / total_action_time if total_action_time > 0 else 0.0
    rsi_std = rsi / (height / 100.0) if not np.isnan(height) else np.nan
    
    # 向心功率峰值
    con_power_peak = event_vals["最大向心功率"]["P"]
    con_power_peak_std = con_power_peak / bw_n
    
    # 向心做功量
    work_con = np.sum(p_series[lowest_idx:takeoff_idx] * dt)
    work_con_std = work_con / bw_n
    
    # --- 3. 離心牽張肌力特徵 ---
    # 下蹲力峰值 (淨力)
    unweight_force_peak = event_vals["最大失重"]["F"]
    unweight_force_peak_std = unweight_force_peak / bw_n
    
    # 下蹲發力率 (RFD)
    rfd_unweight = unweight_force_peak / (t_unweight - t_som) if t_unweight != t_som else 0.0
    rfd_unweight_std = rfd_unweight / bw_n
    
    # 制動末力值 (重心最低淨力)
    braking_end_force = event_vals["重心最低"]["F"]
    braking_end_force_std = braking_end_force / bw_n
    
    # 制動發力率 (RFD)
    rfd_braking = braking_end_force / (t_lowest - t_braking) if t_lowest != t_braking else 0.0
    rfd_braking_std = rfd_braking / bw_n
    
    # 離心功率峰值
    ecc_power_peak = event_vals["最大離心功率"]["P"]
    ecc_power_peak_std = ecc_power_peak / bw_n
    
    # 離心做功量
    work_ecc = np.sum(p_series[som_idx:lowest_idx] * dt)
    work_ecc_std = work_ecc / bw_n
    
    # --- 4. 動作時宜與下肢勁度特徵 ---
    # 時間期 (Seconds)
    t_unweight_phase = t_braking - t_som
    t_braking_phase = t_lowest - t_braking
    t_prop_phase = t_takeoff - t_lowest
    t_amort_phase = t_amort_end - t_amort_start
    t_total_action = total_action_time
    
    # 時間百分比 (%標準化)
    t_unweight_phase_pct = t_unweight_phase / t_total_action * 100 if t_total_action > 0 else 0.0
    t_braking_phase_pct = t_braking_phase / t_total_action * 100 if t_total_action > 0 else 0.0
    t_prop_phase_pct = t_prop_phase / t_total_action * 100 if t_total_action > 0 else 0.0
    t_amort_phase_pct = t_amort_phase / t_total_action * 100 if t_total_action > 0 else 0.0
    t_total_action_pct = 100.0
    
    # 下肢勁度 (K, N/m)
    s_lowest = abs(event_vals["重心最低"]["S"])
    stiffness = braking_end_force / s_lowest if s_lowest > 0 else 0.0
    stiffness_std = (stiffness / bw_n) * (height / 100.0) if (not np.isnan(height) and bw_n > 0) else np.nan
    
    # --- 5. 輸出普通一維 DataFrame (由寫入端進行 openpyxl 格式化後處理) ---
    headers = [
        "資料", "身高", "體重", "年齡", "衝量",
        "評估指標", "跳躍高度", "推蹬力峰值", "推蹬發力率", "反應力指數", "向心功率峰值", "向心做功量",
        "評估指標", "下蹲力峰值", "下蹲發力率", "制動末力值", "制動發力率", "離心功率峰值", "離心做功量",
        "評估指標", "下蹲期時間", "制動期時間", "推蹬期時間", "攤還期時間", "總動作時間", "下肢勁度"
    ]
    
    raw_row = [
        "原始", height, weight_kg, age, round(impulse_raw, 2),
        "原始", round(jump_height, 2), round(peak_force_con, 2), round(rfd_con, 2), round(rsi, 2), round(con_power_peak, 2), round(work_con, 2),
        "原始", round(unweight_force_peak, 2), round(rfd_unweight, 2), round(braking_end_force, 2), round(rfd_braking, 2), round(ecc_power_peak, 2), round(work_ecc, 2),
        "原始", round(t_unweight_phase, 2), round(t_braking_phase, 2), round(t_prop_phase, 2), round(t_amort_phase, 2), round(t_total_action, 2), round(stiffness, 2)
    ]
    
    std_row = [
        "%標準化", height, weight_kg, age, round(impulse_std, 2) if not np.isnan(impulse_std) else np.nan,
        "%標準化", round(jump_height_std, 2) if not np.isnan(jump_height_std) else np.nan,
        round(peak_force_con_std, 2) if not np.isnan(peak_force_con_std) else np.nan,
        round(rfd_con_std, 2) if not np.isnan(rfd_con_std) else np.nan,
        round(rsi_std, 2) if not np.isnan(rsi_std) else np.nan,
        round(con_power_peak_std, 2) if not np.isnan(con_power_peak_std) else np.nan,
        round(work_con_std, 2) if not np.isnan(work_con_std) else np.nan,
        "%標準化", round(unweight_force_peak_std, 2) if not np.isnan(unweight_force_peak_std) else np.nan,
        round(rfd_unweight_std, 2) if not np.isnan(rfd_unweight_std) else np.nan,
        round(braking_end_force_std, 2) if not np.isnan(braking_end_force_std) else np.nan,
        round(rfd_braking_std, 2) if not np.isnan(rfd_braking_std) else np.nan,
        round(ecc_power_peak_std, 2) if not np.isnan(ecc_power_peak_std) else np.nan,
        round(work_ecc_std, 2) if not np.isnan(work_ecc_std) else np.nan,
        "%標準化", round(t_unweight_phase_pct, 2), round(t_braking_phase_pct, 2), round(t_prop_phase_pct, 2), round(t_amort_phase_pct, 2), round(t_total_action_pct, 2), round(stiffness_std, 2) if not np.isnan(stiffness_std) else np.nan
    ]
    
    df_cmj = pd.DataFrame([raw_row, std_row], columns=headers)
    return df_cmj

def format_cmj_sheet(file_path):
    """
    使用 openpyxl 載入寫好的 Excel 檔案，對 CMJ 工作表進行後處理：
    1. 在 Row 1 插入大分類標題
    2. 合併對應的儲存格，以配合實驗室金標準的視覺化結構
    """
    wb = openpyxl.load_workbook(file_path)
    if "CMJ" not in wb.sheetnames:
        wb.close()
        return
        
    ws = wb["CMJ"]
    
    # 插入第一行
    ws.insert_rows(1)
    
    # 寫入大標題名稱
    ws["A1"] = "受試者基本資料"
    ws["F1"] = "下肢動態肌力特徵"
    ws["M1"] = "離心牽張肌力特徵"
    ws["T1"] = "動作時宜與下肢勁度特徵"
    
    # 合併儲存格
    ws.merge_cells("A1:E1")
    ws.merge_cells("F1:L1")
    ws.merge_cells("M1:S1")
    ws.merge_cells("T1:Z1")
    
    wb.save(file_path)
    wb.close()
